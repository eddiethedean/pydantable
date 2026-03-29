"""Tier-2/3 formats and SDK bridges (optional extras; many are **experimental**)."""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from typing import Any, BinaryIO, Iterator, TextIO, cast

from .batches import ensure_rectangular

_EXPERIMENTAL_ENV = "PYDANTABLE_IO_EXPERIMENTAL"


def _require_experimental(experimental: bool, feature: str) -> None:
    if experimental:
        return
    if os.environ.get(_EXPERIMENTAL_ENV, "").lower() in ("1", "true", "yes"):
        return
    raise ValueError(
        f"{feature} is experimental. Pass experimental=True or set {_EXPERIMENTAL_ENV}=1."
    )


def read_excel(
    path: str | Path,
    *,
    sheet_name: str | int = 0,
    experimental: bool = True,
) -> dict[str, list[Any]]:
    """Load the first sheet (or ``sheet_name``) from ``.xlsx`` via openpyxl → ``dict[str, list]``."""
    _require_experimental(experimental, "Excel ingestion")
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "read_excel requires openpyxl (pip install 'pydantable[excel]')."
        ) from e
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    out: dict[str, list[Any]] = {h: [] for h in header}
    for row in rows[1:]:
        for i, h in enumerate(header):
            out[h].append(row[i] if i < len(row) else None)
    return out


def iter_excel(
    path: str | Path,
    *,
    sheet_name: str | int = 0,
    batch_size: int = 65_536,
    experimental: bool = True,
) -> Iterator[dict[str, list[Any]]]:
    """Yield Excel rows as ``dict[str, list]`` batches (openpyxl read-only)."""
    _require_experimental(experimental, "Excel ingestion")
    if batch_size <= 0:
        raise ValueError("batch_size must be a positive integer")
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "iter_excel requires openpyxl (pip install 'pydantable[excel]')."
        ) from e
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
    try:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first = next(rows_iter)
        except StopIteration:
            return
        header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(first)]
        out: dict[str, list[Any]] = {h: [] for h in header}
        n = 0
        for row in rows_iter:
            for i, h in enumerate(header):
                out[h].append(row[i] if i < len(row) else None)
            n += 1
            if n >= batch_size:
                ensure_rectangular(out)
                yield out
                out = {h: [] for h in header}
                n = 0
        if n:
            ensure_rectangular(out)
            yield out
    finally:
        wb.close()


def read_delta(
    path: str | Path,
    *,
    experimental: bool = True,
) -> dict[str, list[Any]]:
    """Read a Delta table directory via PyArrow dataset (``[arrow]`` extra)."""
    _require_experimental(experimental, "Delta Lake ingestion")
    try:
        import pyarrow.dataset as ds  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "read_delta requires pyarrow with dataset support (pip install 'pydantable[arrow]')."
        ) from e
    from .arrow import arrow_table_to_column_dict

    dset = ds.dataset(path, format="parquet")
    table = dset.to_table()
    return arrow_table_to_column_dict(table)


def iter_delta(
    path: str | Path,
    *,
    batch_size: int = 65_536,
    experimental: bool = True,
) -> Iterator[dict[str, list[Any]]]:
    """Yield Delta (Parquet dataset) batches via PyArrow dataset scanner."""
    _require_experimental(experimental, "Delta Lake ingestion")
    if batch_size <= 0:
        raise ValueError("batch_size must be a positive integer")
    try:
        import pyarrow.dataset as ds  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "iter_delta requires pyarrow.dataset (pip install 'pydantable[arrow]')."
        ) from e
    dset = ds.dataset(path, format="parquet")
    for record_batch in dset.to_batches(batch_size=batch_size):
        d = record_batch.to_pydict()
        out = {k: list(v) for k, v in d.items()}
        ensure_rectangular(out)
        yield out


def read_avro(
    path: str | Path,
    *,
    experimental: bool = True,
) -> dict[str, list[Any]]:
    _require_experimental(experimental, "Avro ingestion")
    try:
        import pyarrow as pa  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "read_avro requires pyarrow (pip install 'pydantable[arrow]')."
        ) from e
    from .arrow import arrow_table_to_column_dict

    try:
        table = pa.avro.read_table(str(path))
    except AttributeError as e:
        raise ImportError("pyarrow.avro is not available in this pyarrow build.") from e
    return arrow_table_to_column_dict(table)


def iter_avro(
    path: str | Path,
    *,
    batch_size: int = 65_536,
    experimental: bool = True,
) -> Iterator[dict[str, list[Any]]]:
    """Yield Avro batches via PyArrow (falls back to full read if needed)."""
    _require_experimental(experimental, "Avro ingestion")
    if batch_size <= 0:
        raise ValueError("batch_size must be a positive integer")
    try:
        import pyarrow as pa  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "iter_avro requires pyarrow (pip install 'pydantable[arrow]')."
        ) from e
    try:
        reader = pa.avro.open_file(str(path))  # type: ignore[attr-defined]
    except Exception:
        # Some builds lack streaming Avro; fall back to eager.
        yield read_avro(path, experimental=True)
        return
    with reader:
        for rb in reader.iter_batches(batch_size=batch_size):
            d = rb.to_pydict()
            out = {k: list(v) for k, v in d.items()}
            ensure_rectangular(out)
            yield out


def read_orc(
    path: str | Path,
    *,
    experimental: bool = True,
) -> dict[str, list[Any]]:
    _require_experimental(experimental, "ORC ingestion")
    try:
        import pyarrow.orc as orc  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "read_orc requires pyarrow.orc (pip install 'pydantable[arrow]')."
        ) from e
    from .arrow import arrow_table_to_column_dict

    with open(path, "rb") as f:
        table = orc.ORCFile(f).read()
    return arrow_table_to_column_dict(table)


def iter_orc(
    path: str | Path,
    *,
    batch_size: int = 65_536,
    experimental: bool = True,
) -> Iterator[dict[str, list[Any]]]:
    """Yield ORC batches via PyArrow."""
    _require_experimental(experimental, "ORC ingestion")
    if batch_size <= 0:
        raise ValueError("batch_size must be a positive integer")
    try:
        import pyarrow.orc as orc  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "iter_orc requires pyarrow.orc (pip install 'pydantable[arrow]')."
        ) from e
    with open(path, "rb") as f:
        of = orc.ORCFile(f)
        for rb in of.iter_batches(batch_size=batch_size):
            d = rb.to_pydict()
            out = {k: list(v) for k, v in d.items()}
            ensure_rectangular(out)
            yield out


def read_bigquery(
    query: str,
    *,
    project: str | None = None,
    experimental: bool = True,
    **kwargs: Any,
) -> dict[str, list[Any]]:
    """Run a BigQuery SQL string via ``google-cloud-bigquery`` → ``dict[str, list]``."""
    _require_experimental(experimental, "BigQuery ingestion")
    try:
        from google.cloud import bigquery  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "read_bigquery requires google-cloud-bigquery (pip install 'pydantable[bq]')."
        ) from e
    from .arrow import arrow_table_to_column_dict

    client = bigquery.Client(project=project, **kwargs)
    rows = client.query(query).result()
    table = rows.to_arrow()
    return arrow_table_to_column_dict(table)


def iter_bigquery(
    query: str,
    *,
    project: str | None = None,
    batch_size: int = 65_536,
    experimental: bool = True,
    **kwargs: Any,
) -> Iterator[dict[str, list[Any]]]:
    """Yield BigQuery results in Arrow-backed batches when available."""
    _require_experimental(experimental, "BigQuery ingestion")
    if batch_size <= 0:
        raise ValueError("batch_size must be a positive integer")
    try:
        from google.cloud import bigquery  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "iter_bigquery requires google-cloud-bigquery (pip install 'pydantable[bq]')."
        ) from e
    client = bigquery.Client(project=project, **kwargs)
    rows = client.query(query).result(page_size=batch_size)
    # Prefer Arrow streaming if supported by the client version.
    if hasattr(rows, "to_arrow_iterable"):
        for rb in rows.to_arrow_iterable():  # type: ignore[attr-defined]
            d = rb.to_pydict()
            out = {k: list(v) for k, v in d.items()}
            ensure_rectangular(out)
            yield out
        return
    # Fallback: materialize Arrow table then chunk.
    yield read_bigquery(query, project=project, experimental=True, **kwargs)


def read_snowflake(
    sql: str,
    *,
    experimental: bool = True,
    **connect_kwargs: Any,
) -> dict[str, list[Any]]:
    """Execute ``sql`` on Snowflake via ``snowflake-connector-python`` (experimental)."""
    _require_experimental(experimental, "Snowflake ingestion")
    try:
        import snowflake.connector  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "read_snowflake requires snowflake-connector-python "
            "(pip install 'pydantable[snowflake]')."
        ) from e
    conn = snowflake.connector.connect(**connect_kwargs)
    try:
        cur = conn.cursor()
        cur.execute(sql)
        cols = [c[0] for c in cur.description or []]
        fetched = cur.fetchall()
        if not cols:
            return {}
        return {cols[i]: [row[i] for row in fetched] for i in range(len(cols))}
    finally:
        conn.close()


def iter_snowflake(
    sql: str,
    *,
    batch_size: int = 65_536,
    experimental: bool = True,
    **connect_kwargs: Any,
) -> Iterator[dict[str, list[Any]]]:
    """Yield Snowflake query results in batches (cursor.fetchmany)."""
    _require_experimental(experimental, "Snowflake ingestion")
    if batch_size <= 0:
        raise ValueError("batch_size must be a positive integer")
    try:
        import snowflake.connector  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "iter_snowflake requires snowflake-connector-python "
            "(pip install 'pydantable[snowflake]')."
        ) from e
    conn = snowflake.connector.connect(**connect_kwargs)
    try:
        cur = conn.cursor()
        cur.execute(sql)
        cols = [c[0] for c in cur.description or []]
        if not cols:
            return
        while True:
            chunk = cur.fetchmany(batch_size)
            if not chunk:
                break
            out = {cols[i]: [row[i] for row in chunk] for i in range(len(cols))}
            ensure_rectangular(out)
            yield out
    finally:
        conn.close()


def read_kafka_json_batch(
    topic: str,
    *,
    bootstrap_servers: str,
    max_messages: int = 100,
    experimental: bool = True,
    **consumer_config: Any,
) -> dict[str, list[Any]]:
    """
    Poll JSON payloads from ``topic`` into columns ``key``, ``value``, ``partition``, ``offset``.

    **At-least-once** delivery only; values must be JSON objects whose keys become columns when
    unioning (best-effort flatten).
    """
    _require_experimental(experimental, "Kafka ingestion")
    try:
        from kafka import KafkaConsumer  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "read_kafka_json_batch requires kafka-python (pip install 'pydantable[kafka]')."
        ) from e
    consumer = KafkaConsumer(
        topic,
        bootstrap_servers=bootstrap_servers,
        value_deserializer=lambda b: json.loads(b.decode("utf-8")),
        **consumer_config,
    )
    rows: list[dict[str, Any]] = []
    try:
        for _ in range(max_messages):
            pack = consumer.poll(timeout_ms=2000)
            if not pack:
                break
            for _tp, messages in pack.items():
                for msg in messages:
                    val = msg.value if isinstance(msg.value, dict) else {}
                    row = {
                        "key": msg.key.decode("utf-8") if msg.key else None,
                        "partition": msg.partition,
                        "offset": msg.offset,
                        **val,
                    }
                    rows.append(row)
                    if len(rows) >= max_messages:
                        break
                if len(rows) >= max_messages:
                    break
            if len(rows) >= max_messages:
                break
    finally:
        consumer.close()
    if not rows:
        return {}
    keys = sorted({k for r in rows for k in r})
    return {k: [r.get(k) for r in rows] for k in keys}


def iter_kafka_json(
    topic: str,
    *,
    bootstrap_servers: str,
    max_messages: int | None = None,
    batch_size: int = 1000,
    experimental: bool = True,
    **consumer_config: Any,
) -> Iterator[dict[str, list[Any]]]:
    """
    Stream JSON payloads from Kafka, yielding batches as ``dict[str, list]``.

    Stops after `max_messages` if provided; otherwise runs until poll returns empty.
    """
    _require_experimental(experimental, "Kafka ingestion")
    if batch_size <= 0:
        raise ValueError("batch_size must be a positive integer")
    try:
        from kafka import KafkaConsumer  # type: ignore[import-not-found]
    except ImportError as e:
        raise ImportError(
            "iter_kafka_json requires kafka-python (pip install 'pydantable[kafka]')."
        ) from e
    consumer = KafkaConsumer(
        topic,
        bootstrap_servers=bootstrap_servers,
        value_deserializer=lambda b: json.loads(b.decode("utf-8")),
        **consumer_config,
    )
    seen = 0
    rows: list[dict[str, Any]] = []
    try:
        while True:
            if max_messages is not None and seen >= max_messages:
                break
            pack = consumer.poll(timeout_ms=2000)
            if not pack:
                break
            for _tp, messages in pack.items():
                for msg in messages:
                    val = msg.value if isinstance(msg.value, dict) else {}
                    row = {
                        "key": msg.key.decode("utf-8") if msg.key else None,
                        "partition": msg.partition,
                        "offset": msg.offset,
                        **val,
                    }
                    rows.append(row)
                    seen += 1
                    if len(rows) >= batch_size:
                        keys = sorted({k for r in rows for k in r})
                        out = {k: [r.get(k) for r in rows] for k in keys}
                        ensure_rectangular(out)
                        yield out
                        rows = []
                    if max_messages is not None and seen >= max_messages:
                        break
                if max_messages is not None and seen >= max_messages:
                    break
    finally:
        consumer.close()
    if rows:
        keys = sorted({k for r in rows for k in r})
        out = {k: [r.get(k) for r in rows] for k in keys}
        ensure_rectangular(out)
        yield out

def read_csv_stdin(
    stream: TextIO | None = None,
    *,
    engine: str = "auto",
) -> dict[str, list[Any]]:
    """Read CSV from ``stdin`` (or ``stream``) via a temporary file + :func:`materialize_csv`."""
    import tempfile

    from . import materialize_csv

    fh = stream or sys.stdin
    data = fh.read()
    raw = data.encode("utf-8") if isinstance(data, str) else data
    fd, name = tempfile.mkstemp(suffix=".csv")
    os.close(fd)
    path = Path(name)
    try:
        path.write_bytes(raw)
        return materialize_csv(str(path), engine=engine)
    finally:
        path.unlink(missing_ok=True)


def write_csv_stdout(
    data: dict[str, list[Any]],
    stream: TextIO | BinaryIO | None = None,
    *,
    engine: str = "auto",
) -> None:
    """Write ``data`` as CSV to ``stdout`` (or ``stream``) using :func:`export_csv` to a temp file."""
    import tempfile

    from . import export_csv

    fd, name = tempfile.mkstemp(suffix=".csv")
    os.close(fd)
    path = Path(name)
    try:
        export_csv(str(path), data, engine=engine)
        out = path.read_bytes()
        if stream is None:
            sys.stdout.buffer.write(out)
        elif hasattr(stream, "buffer"):
            stream.buffer.write(out)
        else:
            cast("TextIO", stream).write(out.decode("utf-8"))
    finally:
        path.unlink(missing_ok=True)
