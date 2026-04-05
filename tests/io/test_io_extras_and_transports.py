"""Tests for optional I/O extras, object-store URIs, stdin/stdout, and mocked cloud SDK paths."""

from __future__ import annotations

import io
import sys
import types
from pathlib import Path
from unittest.mock import patch

import pytest
from pydantable.io import (
    read_csv_stdin,
    read_from_object_store,
    write_csv_stdout,
)
from pydantable.io.extras import (
    iter_delta,
    iter_excel,
    read_avro,
    read_bigquery,
    read_delta,
    read_excel,
    read_kafka_json_batch,
    read_orc,
    read_snowflake,
)
from pydantable.io.rap_support import aread_csv_rap, rap_csv_available


def test_read_excel_roundtrip(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "book.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "name"])
    ws.append([1, "alice"])
    ws.append([2, "bob"])
    wb.save(path)

    got = read_excel(path, experimental=True)
    assert got["id"] == [1, 2]
    assert got["name"] == ["alice", "bob"]


def test_read_excel_requires_experimental(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "e.xlsx"
    Workbook().save(path)
    with pytest.raises(ValueError, match="experimental"):
        read_excel(path, experimental=False)


def test_read_excel_header_only_no_data_rows(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "header_only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["only_header"])
    wb.save(path)
    got = read_excel(path, experimental=True)
    assert got == {"only_header": []}


def test_read_delta_parquet_directory(tmp_path: Path) -> None:
    pytest.importorskip("pyarrow")
    import pyarrow as pa
    import pyarrow.parquet as pq

    root = tmp_path / "lake"
    root.mkdir()
    pq.write_table(pa.table({"v": [10, 20], "w": ["a", "b"]}), root / "part.parquet")

    got = read_delta(root, experimental=True)
    assert got["v"] == [10, 20]
    assert got["w"] == ["a", "b"]


def test_read_delta_requires_experimental(tmp_path: Path) -> None:
    pytest.importorskip("pyarrow")
    root = tmp_path / "d"
    root.mkdir()
    with pytest.raises(ValueError, match="experimental"):
        read_delta(root, experimental=False)


def test_read_from_object_store_file_uri_parquet(tmp_path: Path) -> None:
    pytest.importorskip("fsspec")
    pytest.importorskip("pyarrow")
    import pyarrow as pa
    import pyarrow.parquet as pq

    path = tmp_path / "obj.parquet"
    pq.write_table(pa.table({"z": [7, 8]}), path)
    uri = path.resolve().as_uri()

    got = read_from_object_store(uri, experimental=True, format="parquet")
    assert got == {"z": [7, 8]}


def test_read_from_object_store_csv_and_ndjson(tmp_path: Path) -> None:
    pytest.importorskip("fsspec")

    csv_path = tmp_path / "remote.csv"
    csv_path.write_text("c,d\n1,2\n", encoding="utf-8")
    got_csv = read_from_object_store(
        csv_path.resolve().as_uri(), experimental=True, format="csv"
    )
    assert got_csv["c"] in ([1], ["1"])
    assert got_csv["d"] in ([2], ["2"])

    nd_path = tmp_path / "remote.ndjson"
    nd_path.write_text('{"m": 1}\n{"m": 2}\n', encoding="utf-8")
    got_nd = read_from_object_store(
        nd_path.resolve().as_uri(), experimental=True, format="ndjson"
    )
    assert got_nd == {"m": [1, 2]}


def test_read_from_object_store_jsonl_alias(tmp_path: Path) -> None:
    pytest.importorskip("fsspec")
    nd_path = tmp_path / "lines.jsonl"
    nd_path.write_text('{"k": "u"}\n', encoding="utf-8")
    got = read_from_object_store(
        nd_path.resolve().as_uri(), experimental=True, format="jsonl"
    )
    assert got == {"k": ["u"]}


def test_read_from_object_store_rejects_http() -> None:
    pytest.importorskip("fsspec")
    with pytest.raises(ValueError, match="fetch_parquet_url"):
        read_from_object_store("https://example.com/x.parquet", experimental=True)


def test_read_from_object_store_unsupported_format(tmp_path: Path) -> None:
    pytest.importorskip("fsspec")
    blob = tmp_path / "blob.dat"
    blob.write_bytes(b"x")
    with pytest.raises(ValueError, match="unsupported format"):
        read_from_object_store(blob.resolve().as_uri(), experimental=True, format="xml")


def test_read_from_object_store_experimental_env(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    pytest.importorskip("fsspec")
    pytest.importorskip("pyarrow")
    import pyarrow as pa
    import pyarrow.parquet as pq

    monkeypatch.setenv("PYDANTABLE_IO_EXPERIMENTAL", "1")
    path = tmp_path / "env.parquet"
    pq.write_table(pa.table({"e": [1]}), path)
    got = read_from_object_store(
        path.resolve().as_uri(), experimental=False, format="parquet"
    )
    assert got == {"e": [1]}


def test_read_csv_stdin_via_stringio() -> None:
    stream = io.StringIO("a,b\n9,8\n")
    got = read_csv_stdin(stream, engine="auto")
    assert got["a"] in ([9], ["9"])
    assert got["b"] in ([8], ["8"])


def test_write_csv_stdout_stringio() -> None:
    buf = io.StringIO()
    write_csv_stdout({"x": [1, 2]}, stream=buf, engine="auto")
    raw = buf.getvalue()
    assert "x" in raw
    assert "1" in raw and "2" in raw


@pytest.mark.optional_cloud
def test_read_bigquery_mocked(monkeypatch: pytest.MonkeyPatch) -> None:
    pa = pytest.importorskip("pyarrow")
    pytest.importorskip("google.cloud.bigquery")

    class _Result:
        def to_arrow(self):
            return pa.table({"bq": [11, 22]})

    class _Job:
        def result(self):
            return _Result()

    class _Client:
        def __init__(self, *a: object, **k: object) -> None:
            pass

        def query(self, _q: str):
            return _Job()

    monkeypatch.setattr("google.cloud.bigquery.Client", _Client)
    got = read_bigquery("SELECT 1", experimental=True)
    assert got == {"bq": [11, 22]}


@pytest.mark.optional_cloud
def test_read_snowflake_mocked() -> None:
    pytest.importorskip("snowflake.connector")

    class _Cur:
        description = [("s",), ("n",)]

        def execute(self, _sql: str) -> None:
            return None

        def fetchall(self):
            return [("hi", 3)]

    class _Conn:
        def cursor(self):
            return _Cur()

        def close(self) -> None:
            return None

    with patch("snowflake.connector.connect", return_value=_Conn()):
        got = read_snowflake("SELECT 1", experimental=True, account="a", user="u")
    assert got == {"s": ["hi"], "n": [3]}


def test_read_snowflake_empty_cursor_description() -> None:
    pytest.importorskip("snowflake.connector")

    class _Cur:
        description = None

        def execute(self, _sql: str) -> None:
            return None

        def fetchall(self):
            return []

    class _Conn:
        def cursor(self):
            return _Cur()

        def close(self) -> None:
            return None

    with patch("snowflake.connector.connect", return_value=_Conn()):
        assert (
            read_snowflake("SELECT 1", experimental=True, account="a", user="u") == {}
        )


def test_read_kafka_json_batch_mocked() -> None:
    pytest.importorskip("kafka")

    class _Msg:
        key = b"k"
        value = {"payload": 42}
        partition = 0
        offset = 100

    class _Consumer:
        def __init__(self, *a: object, **k: object) -> None:
            self._done = False

        def poll(self, timeout_ms: int = 0):
            if self._done:
                return {}
            self._done = True
            tp = ("topic", 0)
            return {tp: [_Msg()]}

        def close(self) -> None:
            return None

    with patch("kafka.KafkaConsumer", _Consumer):
        got = read_kafka_json_batch(
            "t1",
            bootstrap_servers="localhost:9092",
            max_messages=5,
            experimental=True,
        )
    assert got["key"] == ["k"]
    assert got["payload"] == [42]
    assert got["partition"] == [0]
    assert got["offset"] == [100]


def test_read_kafka_json_batch_empty_poll() -> None:
    pytest.importorskip("kafka")

    class _Consumer:
        def __init__(self, *a: object, **k: object) -> None:
            pass

        def poll(self, timeout_ms: int = 0):
            return {}

        def close(self) -> None:
            return None

    with patch("kafka.KafkaConsumer", _Consumer):
        assert (
            read_kafka_json_batch(
                "empty",
                bootstrap_servers="localhost:9092",
                max_messages=3,
                experimental=True,
            )
            == {}
        )


def test_read_avro_mocked(tmp_path: Path) -> None:
    pa = pytest.importorskip("pyarrow")

    fake_table = pa.table({"av": ["x"]})
    fake_avro = types.SimpleNamespace(read_table=lambda _p: fake_table)
    path = tmp_path / "f.avro"
    path.write_bytes(b"")
    with patch.object(sys.modules["pyarrow"], "avro", fake_avro, create=True):
        got = read_avro(path, experimental=True)
    assert got == {"av": ["x"]}


def test_read_orc_mocked(tmp_path: Path) -> None:
    pa = pytest.importorskip("pyarrow")

    class _ORC:
        def __init__(self, _f: object) -> None:
            pass

        def read(self):
            return pa.table({"oc": [5]})

    path = tmp_path / "dummy.orc"
    path.write_bytes(b"dummy")
    with patch("pyarrow.orc.ORCFile", _ORC):
        got = read_orc(path, experimental=True)
    assert got == {"oc": [5]}


def test_rap_csv_available_is_bool() -> None:
    assert isinstance(rap_csv_available(), bool)


def _require_rap_open() -> None:
    pytest.importorskip("rapcsv")
    rf = pytest.importorskip("rapfiles")
    if not hasattr(rf, "open"):
        pytest.skip("rapfiles.open not available in this rapfiles build")


@pytest.mark.asyncio
async def test_aread_csv_rap_roundtrip(tmp_path: Path) -> None:
    _require_rap_open()

    path = tmp_path / "rap.csv"
    path.write_text("p,q\n1,2\n3,4\n", encoding="utf-8")
    got = await aread_csv_rap(str(path))
    assert got["p"] == ["1", "3"]
    assert got["q"] == ["2", "4"]


@pytest.mark.asyncio
async def test_aread_csv_rap_header_only_no_data_rows(tmp_path: Path) -> None:
    _require_rap_open()

    # rapcsv may panic on a completely empty file; header-only is the safe "no rows" case.
    path = tmp_path / "headers.csv"
    path.write_text("p,q\n", encoding="utf-8")
    assert await aread_csv_rap(str(path)) == {}


def test_iter_excel_batch_size_invalid(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "b.xlsx"
    Workbook().save(path)
    with pytest.raises(ValueError, match="batch_size"):
        next(iter_excel(path, batch_size=0, experimental=True))


def test_iter_excel_yields_row_batches(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    path = tmp_path / "batched.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    for i in range(5):
        ws.append([i, i * 10])
    wb.save(path)

    batches = list(iter_excel(path, batch_size=2, experimental=True))
    assert len(batches) >= 1
    assert batches[0]["a"] == [0, 1]


def test_iter_delta_yields_batches(tmp_path: Path) -> None:
    pytest.importorskip("pyarrow")
    import pyarrow as pa
    import pyarrow.parquet as pq

    root = tmp_path / "delta_iter"
    root.mkdir()
    pq.write_table(
        pa.table({"u": list(range(20))}),
        root / "p.parquet",
    )
    parts = list(iter_delta(root, batch_size=8, experimental=True))
    assert len(parts) >= 1
    assert sum(len(p["u"]) for p in parts) == 20


def test_read_excel_allows_experimental_false_when_env_set(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    monkeypatch.setenv("PYDANTABLE_IO_EXPERIMENTAL", "1")
    path = tmp_path / "env.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["c"])
    ws.append([1])
    wb.save(path)
    got = read_excel(path, experimental=False)
    assert got["c"] == [1]
